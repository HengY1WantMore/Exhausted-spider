# encoding='utf-8'
import re
from urllib.parse import urlencode
import openpyxl
import xlrd
import requests
from pyquery import PyQuery as pq
from requests import RequestException
from xpinyin import Pinyin


def log(location, text):
    with open(location, "a+", encoding='utf-8') as f:
        f.write(text)


def get_info_excel(file, col):
    res = []
    wb = xlrd.open_workbook(filename=file)  # 打开文件
    sheet = wb.sheet_by_index(0)  # 通过索引获取表格
    col = sheet.col(col)[1:]  # 获取到目标
    for one in col:
        res.append(one.value)
    return res


def write_name(handleName):
    workbook = openpyxl.load_workbook("aim_test.xlsx")
    worksheet = workbook.worksheets[0]
    for x in range(len(handleName)):
        worksheet.cell(x + 2, 3, handleName[x])
    workbook.save(filename="aim_test.xlsx")
    print('名字更新完成')


def write_time(handleTime):
    workbook = openpyxl.load_workbook("aim_test.xlsx")
    worksheet = workbook.worksheets[0]
    for x in range(len(handleName)):
        worksheet.cell(x + 2, 8, handleTime[x])
    workbook.save(filename="aim_test.xlsx")
    print('年份更新完成')


def handleName(english, name):
    for x in range(len(english)):
        if english[x] == '':
            p = Pinyin()
            s = p.get_pinyin(name[x]).split('-')
            english[x] = s[0].capitalize() + ' ' + ''.join(s[1:]).capitalize()
    return english


def handleTime(time):
    for x in range(len(time)):
        if time[x] != '':
            match = '(.*?)年(.*?)月(.*?)日'
            results = re.findall(match, time[x], re.S)
            if results:
                results = list(results[0])
                results[1] = results[1] if len(results[1]) == 2 else '0' + results[1]
                results[2] = results[2] if len(results[2]) == 2 else '0' + results[2]
                time[x] = results[0] + '-' + results[1] + '-' + results[2]
            match = '(.*?)年(.*?)月'
            results = list(re.findall(match, time[x], re.S))
            if results:
                results = list(results[0])
                results[1] = results[1] if len(results[1]) == 2 else '0' + results[1]
                time[x] = results[0] + '-' + results[1] + '-00'
            match = '(.*?)年'
            results = list(re.findall(match, time[x], re.S))
            if results:
                time[x] = results[0] + '-00-00'
    return time


def write(info, col):
    workbook = openpyxl.load_workbook("aim_test.xlsx")
    worksheet = workbook.worksheets[0]
    for x in info:
        worksheet.cell(x + 2, col, '')
    workbook.save(filename="aim_test.xlsx")
    print('更新完成')


def get_page(name):
    data = {
        'keyword': name,
    }
    name = str(urlencode(data)).split('=')[1]
    url = 'https://baike.baidu.com/item/' + name
    header = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.107 Safari/537.36'
    }
    response = requests.get(url, headers=header)
    try:
        if response.status_code == 200:
            return response.text.replace('&nbsp;', '')
    except RequestException:
        print('请求出错')
        return None


def parse_page(page, name):
    doc = pq(page)
    info_left = doc('.basicInfo-block.basicInfo-left').text().replace('\n', ',').split(',')
    group1 = [info_left[i:i + 2] for i in range(0, len(info_left), 2)]
    info_right = doc('.basicInfo-block.basicInfo-right').text().replace('\n', ',').split(',')
    group2 = [info_right[i:i + 2] for i in range(0, len(info_right), 2)]
    abstract = doc('.para').text().replace('\u30fb', '')
    print(f"{name} 信息：{group1} {group2}  总体信息： {abstract}")
    log('doubt.txt', f"{name} 信息：{group1} {group2}  总体信息： {abstract} \n")


if __name__ == '__main__':
    info = get_info_excel('aim_test.xlsx', 49)
    info_young = get_info_excel('aim_test.xlsx', 55)
    name = get_info_excel('aim_test.xlsx', 1)
    time = get_info_excel('aim_test.xlsx', 7)
    # time_list = []
    # country = get_info_excel('aim_test.xlsx', 10)
    # country_list = []
    # for index, info in enumerate(time):
    #     if info == '':
    #         time_list.append(index)
    # for index, info in enumerate(country):
    #     if info != '':
    #         country_list.append(index)
    # total = list(set(time_list + country_list))
    # for x in total:
    #     one = name[x]
    #     page = get_page(one)
    #     parse_page(page, one)
    workbook = openpyxl.load_workbook("aim_test.xlsx")
    worksheet = workbook.worksheets[0]
    for index,info in enumerate(time):
        if info != '':
            info = info.replace('-', '/')
            worksheet.cell(index + 2, 8, info)
    workbook.save(filename="aim_test.xlsx")
    print('更新完成')
